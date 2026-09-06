import sys
import unittest
import base64
import io
from unittest.mock import MagicMock

import pandas as pd
from openpyxl import load_workbook

sys.modules.setdefault("pyodbc", MagicMock())

from engine.exporter import Exporter
from engine.full_reconciliation import FullReconciliationEngine


def engine():
    return FullReconciliationEngine(pd.DataFrame(), pd.DataFrame(), pd.DataFrame())


class FullDispatchResidualTests(unittest.TestCase):
    def test_dispatch_trusts_exact_gtin_already_proven_in_batch_master(self):
        subject = engine()
        sfda = pd.DataFrame([{
            "GTIN": "05014124173224",
            "Drug Name": "FENTANYL 50 MG MCG/ML SOLUTION FOR INJECTION",
            "BN": "0154941", "Expiry Date": "2028-01-31",
            "Quantity": 6186, "Active": 3104,
            "Quantity sent pending": 3082,
            "Quantity Receive Pending": 0,
        }])
        master = pd.DataFrame([{
            "GTIN": "05014124173224", "BN": "0154941",
            "Expiry Month Key": "2028-01",
            "Generic Item Number": "5137230500300",
            "Trade Description": "UNRELATED CURRENT FORMAT",
            "Description": "", "Item Family Group": "PHARMACEUTICALS",
        }])

        conservative = subject.prepare_stage2_sfda_identity(sfda, master)
        dispatch_identity = subject.prepare_stage2_sfda_identity(
            sfda,
            master,
            allow_shared_regulatory_identity=True,
            trust_batch_master_gtin=True,
        )

        self.assertTrue(conservative.empty)
        self.assertEqual(len(dispatch_identity), 1)
        self.assertEqual(dispatch_identity.loc[0, "GTIN"], "05014124173224")
        self.assertEqual(
            dispatch_identity.loc[0, "Generic Item Number"],
            "5137230500300",
        )

    def test_dispatch_keeps_all_generics_with_same_proven_regulatory_gtin(self):
        subject = engine()
        sfda = pd.DataFrame([{
            "GTIN": "06285111000802", "Drug Name": "LIDOCAINE INJECTION",
            "BN": "155320", "Expiry Date": "2028-09-30",
            "Quantity": 11798, "Active": 11542,
            "Quantity sent pending": 256,
            "Quantity Receive Pending": 0,
        }])
        master = pd.DataFrame([
            {
                "GTIN": "06285111000802", "BN": "155320",
                "Expiry Month Key": "2028-09",
                "Generic Item Number": generic,
                "Trade Description": "LIDOCAINE INJECTION",
                "Description": "", "Item Family Group": "PHARMACEUTICALS",
            }
            for generic in ["5114290800600", "5114290802800"]
        ])

        dispatch_identity = subject.prepare_stage2_sfda_identity(
            sfda,
            master,
            allow_shared_regulatory_identity=True,
            trust_batch_master_gtin=True,
        )

        self.assertEqual(len(dispatch_identity), 2)
        self.assertEqual(
            set(dispatch_identity["Generic Item Number"]),
            {"5114290800600", "5114290802800"},
        )

    def test_sfda_inventory_comparison_accepts_sub_pack_difference(self):
        subject = engine()
        sfda = pd.DataFrame([{
            "GTIN": "06281234567890", "Drug Name": "Drug A", "BN": "B001",
            "Expiry Date": "2028-01-15", "Quantity": 1, "Active": 1,
            "Quantity sent pending": 0, "Quantity Receive Pending": 0,
        }])
        inventory = pd.DataFrame([{
            "BN": "B001", "Expiry Date": "2028-01-15", "Available Quantity": 3,
            "Trade Name": "Drug A", "Generic Item Number": "1001",
            "Trade Item Number": "T1",
        }])
        batch_master = pd.DataFrame([{
            "GTIN": "06281234567890", "BN": "B001",
            "Expiry Month Key": "2028-01", "Generic Item Number": "1001",
            "PackageSize": 4, "Total Dispatched Qty Pack": 0,
        }])

        comparison = subject.build_sfda_inventory_comparison(
            sfda, inventory, batch_master
        )

        self.assertEqual(comparison.loc[0, "Matching Status"], "MATCH")
        self.assertEqual(
            float(comparison.loc[0, "Current Inventory Quantity Pack"]), 0.75
        )
        output = Exporter.build_full_reconciliation_summary_workbook(
            pd.DataFrame([{"Metric": "Matching percentage", "Value": 100}]),
            comparison,
        )
        payload = output["Full_Reconciliation_Summary.xlsx"]["content"]
        workbook = load_workbook(io.BytesIO(base64.b64decode(payload)))
        self.assertEqual(workbook.sheetnames, ["Summary", "SFDA vs Inventory"])

    def test_inventory_difference_uses_customer_history_then_dummy_gln(self):
        subject = engine()
        inventory = pd.DataFrame([
            {
                "BN": "B001",
                "Expiry Date": "2027-11-15",
                "Available Quantity": 100,
                "Trade Name": "Drug A",
                "Generic Item Number": "1001",
                "Trade Item Number": "INVENTORY-T1",
            }
        ])
        validated_sfda = pd.DataFrame([
            {
                "GTIN": "06281234567890",
                "Drug Name": "Drug A",
                "BN": "B001",
                "Expiry Date": pd.Timestamp("2027-11-15"),
                "Expiry Month Key": "2027-11",
                "Generic Item Number": "1001",
                "Quantity": 500,
                "Active": 500,
                "Quantity sent pending": 0,
                "Quantity Receive Pending": 0,
            }
        ])
        customer_history = pd.DataFrame([
            {
                "To Address": "Customer A",
                "GLN": "99999999999999",
                "GTIN": "06281234567890",
                "Drug Name": "Drug A",
                "Generic Item Number": "1001",
                "Trade Description": "Drug A",
                "BN": "B001",
                "Expiry Date": pd.Timestamp("2027-11-15"),
                "Expiry Month Key": "2027-11",
                "PackageSize": 1,
                "Dispatch Quantity Each": 50,
                "Dispatch Quantity Pack": 50,
                "First Dispatch Date": pd.Timestamp("2026-06-01"),
                "Last Dispatch Date": pd.Timestamp("2026-06-01"),
                "Custody": "",
                "Trade Item Number": "DISPATCH-T1",
            }
        ])
        batch_master = pd.DataFrame([
            {
                "BN": "B001",
                "Expiry Month Key": "2027-11",
                "Generic Item Number": "1001",
                "Trade Item Number": "MASTER-T1",
            }
        ])

        result = subject.build_dispatch_reconciliation(
            inventory,
            pd.DataFrame(),
            customer_history,
            confirmed_full_dispatch_df=pd.DataFrame(),
            batch_master_df=batch_master,
            validated_sfda_identity_df=validated_sfda,
        )

        self.assertEqual(int(result["To Be Dispatch"].sum()), 400)
        customer_row = result.loc[result["To Address"].eq("Customer A")].iloc[0]
        dummy_row = result.loc[
            result["To Address"].eq("UNMAPPED / DUMMY GLN")
        ].iloc[0]
        self.assertEqual(int(customer_row["To Be Dispatch"]), 50)
        self.assertEqual(customer_row["Trade Item Number"], "DISPATCH-T1")
        self.assertEqual(int(dummy_row["To Be Dispatch"]), 350)
        self.assertEqual(dummy_row["GLN"], "99999999999999")
        self.assertEqual(dummy_row["Trade Item Number"], "MASTER-T1")

        files = Exporter.build_dispatch_files_by_customer(result)
        self.assertEqual(list(files), ["99999999999999_001.csv"])
        self.assertIn(
            "06281234567890;400;B001;15-11-2027",
            next(iter(files.values())),
        )

    def test_shared_gtin_batch_is_dispatched_once_across_multiple_generics(self):
        subject = engine()
        inventory = pd.DataFrame([
            {
                "BN": "B-MULTI", "Expiry Date": "2028-09-15",
                "Available Quantity": 60, "Trade Name": "Drug Multi",
                "Generic Item Number": "G1", "Trade Item Number": "T1",
            },
            {
                "BN": "B-MULTI", "Expiry Date": "2028-09-15",
                "Available Quantity": 40, "Trade Name": "Drug Multi",
                "Generic Item Number": "G2", "Trade Item Number": "T2",
            },
        ])
        validated_sfda = pd.DataFrame([
            {
                "GTIN": "06281234560001", "Drug Name": "Drug Multi",
                "BN": "B-MULTI", "Expiry Date": pd.Timestamp("2028-09-15"),
                "Expiry Month Key": "2028-09", "Generic Item Number": generic,
                "Quantity": 200, "Active": 200,
                "Quantity sent pending": 0, "Quantity Receive Pending": 0,
            }
            for generic in ["G1", "G2"]
        ])
        customer_history = pd.DataFrame([
            {
                "To Address": f"Customer {generic}", "GLN": gln,
                "GTIN": "06281234560001", "Drug Name": "Drug Multi",
                "Generic Item Number": generic, "Trade Description": "Drug Multi",
                "BN": "B-MULTI", "Expiry Date": pd.Timestamp("2028-09-15"),
                "Expiry Month Key": "2028-09", "PackageSize": 1,
                "Dispatch Quantity Each": quantity,
                "Dispatch Quantity Pack": quantity,
                "First Dispatch Date": pd.Timestamp("2028-01-01"),
                "Last Dispatch Date": pd.Timestamp("2028-01-01"),
                "Custody": "", "Trade Item Number": trade,
            }
            for generic, gln, quantity, trade in [
                ("G1", "1000000000001", 30, "T1"),
                ("G2", "1000000000002", 20, "T2"),
            ]
        ])
        batch_master = pd.DataFrame([
            {
                "BN": "B-MULTI", "Expiry Month Key": "2028-09",
                "Generic Item Number": generic, "Trade Item Number": trade,
                "PackageSize": 1,
            }
            for generic, trade in [("G1", "T1"), ("G2", "T2")]
        ])

        result = subject.build_dispatch_reconciliation(
            inventory, pd.DataFrame(), customer_history,
            confirmed_full_dispatch_df=pd.DataFrame(),
            batch_master_df=batch_master,
            validated_sfda_identity_df=validated_sfda,
        )

        self.assertEqual(int(result["To Be Dispatch"].sum()), 100)
        self.assertEqual(int(result.loc[
            result["To Address"].eq("UNMAPPED / DUMMY GLN"),
            "To Be Dispatch",
        ].sum()), 50)

    def test_regulatory_difference_without_customer_history_uses_dummy_gln(self):
        subject = engine()
        inventory = pd.DataFrame([{
            "BN": "B-DUMMY", "Expiry Date": "2029-03-15",
            "Available Quantity": 7500, "Trade Name": "Drug Dummy",
            "Generic Item Number": "G1", "Trade Item Number": "T1",
        }])
        validated_sfda = pd.DataFrame([{
            "GTIN": "06281234560002", "Drug Name": "Drug Dummy",
            "BN": "B-DUMMY", "Expiry Date": pd.Timestamp("2029-03-15"),
            "Expiry Month Key": "2029-03", "Generic Item Number": "G1",
            "Quantity": 375, "Active": 375,
            "Quantity sent pending": 0, "Quantity Receive Pending": 0,
        }])
        batch_master = pd.DataFrame([{
            "BN": "B-DUMMY", "Expiry Month Key": "2029-03",
            "Generic Item Number": "G1", "Trade Item Number": "T1",
            "PackageSize": 40,
        }])

        result = subject.build_dispatch_reconciliation(
            inventory, pd.DataFrame(), pd.DataFrame(),
            confirmed_full_dispatch_df=pd.DataFrame(),
            batch_master_df=batch_master,
            validated_sfda_identity_df=validated_sfda,
        )

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "GLN"], "99999999999999")
        self.assertEqual(int(result.loc[0, "To Be Dispatch"]), 187)

    def test_full_dispatch_report_schema_includes_trade_item_number(self):
        self.assertIn(
            "Trade Item Number",
            FullReconciliationEngine.DISPATCH_RECONCILIATION_COLUMNS,
        )
        self.assertIn(
            "Trade Item Number",
            Exporter.FULL_DISPATCH_RECONCILIATION_COLUMNS,
        )
        for duplicate_column in [
            "Reserved Full Dispatch Quantity Each",
            "Reserved Full Dispatch Quantity Pack",
        ]:
            self.assertNotIn(
                duplicate_column,
                FullReconciliationEngine.DISPATCH_RECONCILIATION_COLUMNS,
            )
            self.assertNotIn(
                duplicate_column,
                Exporter.FULL_DISPATCH_RECONCILIATION_COLUMNS,
            )

    def test_cutover_closes_legacy_customer_history_and_reserves_new_growth(self):
        subject = engine()
        inventory = pd.DataFrame([{
            "BN": "B001", "Expiry Date": "2027-11-15",
            "Available Quantity": 70, "Trade Name": "Drug A",
            "Generic Item Number": "1001", "Trade Item Number": "T1",
        }])
        validated_sfda = pd.DataFrame([{
            "GTIN": "06281234567890", "Drug Name": "Drug A", "BN": "B001",
            "Expiry Date": pd.Timestamp("2027-11-15"),
            "Expiry Month Key": "2027-11", "Generic Item Number": "1001",
            "Quantity": 100, "Active": 100, "Quantity sent pending": 0,
            "Quantity Receive Pending": 0,
        }])
        customer_history = pd.DataFrame([{
            "To Address": "Customer A", "GLN": "1000000000001",
            "GTIN": "06281234567890", "Drug Name": "Drug A",
            "Generic Item Number": "1001", "Trade Description": "Drug A",
            "BN": "B001", "Expiry Date": pd.Timestamp("2027-11-15"),
            "Expiry Month Key": "2027-11", "PackageSize": 1,
            "Dispatch Quantity Each": 80, "Dispatch Quantity Pack": 80,
            "First Dispatch Date": pd.Timestamp("2026-06-01"),
            "Last Dispatch Date": pd.Timestamp("2026-09-01"), "Custody": "",
            "Trade Item Number": "T1",
        }])
        cutover = pd.DataFrame([{
            "BN": "B001", "Expiry Month Key": "2027-11",
            "Generic Item Number": "1001", "To Address": "Customer A",
            "GLN": "1000000000001", "Cutover Closed Quantity Each": 50,
            "Cutover Closed Quantity Pack": 50,
        }])
        reserved = pd.DataFrame([{
            "BN": "B001", "Expiry Date": pd.Timestamp("2027-11-15"),
            "Generic Item Number": "1001", "To Address": "Customer A",
            "GLN": "1000000000001",
            "Reserved Full Dispatch Quantity Each": 10,
            "Reserved Full Dispatch Quantity Pack": 10,
            "Confirmed Full Dispatch Quantity Each": 0,
            "Confirmed Full Dispatch Quantity Pack": 0,
        }])

        result = subject.build_dispatch_reconciliation(
            inventory,
            pd.DataFrame(),
            customer_history,
            confirmed_full_dispatch_df=reserved,
            batch_master_df=pd.DataFrame([{
                "BN": "B001", "Expiry Month Key": "2027-11",
                "Generic Item Number": "1001", "Trade Item Number": "T1",
            }]),
            validated_sfda_identity_df=validated_sfda,
            cutover_baseline_df=cutover,
        )

        customer = result.loc[result["To Address"].eq("Customer A")].iloc[0]
        self.assertEqual(int(customer["Historical Dispatch Quantity Pack"]), 80)
        self.assertEqual(int(customer["Available Historical Dispatch Quantity Pack"]), 20)
        self.assertEqual(int(customer["To Be Dispatch"]), 20)
        self.assertEqual(int(result["To Be Dispatch"].sum()), 20)

        # Once the complete 30-pack post-cutover movement is reserved, an
        # unchanged SFDA report must not regenerate it for the customer or the
        # residual Dummy GLN.
        reserved.loc[:, "Reserved Full Dispatch Quantity Each"] = 30
        reserved.loc[:, "Reserved Full Dispatch Quantity Pack"] = 30
        rerun = subject.build_dispatch_reconciliation(
            inventory,
            pd.DataFrame(),
            customer_history,
            confirmed_full_dispatch_df=reserved,
            batch_master_df=pd.DataFrame([{
                "BN": "B001", "Expiry Month Key": "2027-11",
                "Generic Item Number": "1001", "Trade Item Number": "T1",
            }]),
            validated_sfda_identity_df=validated_sfda,
            cutover_baseline_df=cutover,
        )
        self.assertEqual(int(rerun["To Be Dispatch"].sum()), 0)

if __name__ == "__main__":
    unittest.main()
