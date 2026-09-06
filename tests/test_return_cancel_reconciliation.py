import base64
import io
import sys
import unittest
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import load_workbook

# The pure reconciliation tests do not open SQL connections. Production installs
# pyodbc from requirements.txt; the lightweight local test runtime does not.
sys.modules.setdefault("pyodbc", MagicMock())

from engine.exporter import Exporter
from engine.full_reconciliation import FullReconciliationEngine
from engine.normalizer import Normalizer
from engine.inbound_classification import (
    CUSTOMER_RETURN,
    STO_RETURN,
    classify_inbound_shipment,
)


def engine():
    return FullReconciliationEngine(pd.DataFrame(), pd.DataFrame(), pd.DataFrame())


def customer_history(quantity_each=100, package_size=10, customer_code=""):
    return pd.DataFrame([
        {
            "To Address": "Customer A - Riyadh",
            "Customer Code": customer_code,
            "GLN": "6281234567890",
            "GTIN": "06281234567890",
            "Drug Name": "Drug A",
            "Generic Item Number": "1001",
            "Trade Description": "Drug A",
            "BN": "B001",
            "Expiry Date": pd.Timestamp("2027-08-31"),
            "Expiry Month Key": "2027-08",
            "PackageSize": package_size,
            "Dispatch Quantity Each": quantity_each,
            "Dispatch Quantity Pack": quantity_each / package_size,
            "First Dispatch Date": pd.Timestamp("2026-07-01"),
            "Last Dispatch Date": pd.Timestamp("2026-07-01"),
            "Custody": "",
            "Trade Item Number": "1001",
        }
    ])


def returns_history(
    quantity_each=30,
    return_from="customer a riyadh",
    return_type=CUSTOMER_RETURN,
    return_from_code="",
    package_size=10,
):
    return pd.DataFrame([
        {
            "Return Type": return_type,
            "Inbound Shipment": "TRK30-0001" if return_type == CUSTOMER_RETURN else "TRK49-0001",
            "Return From": return_from,
            "Return From Code": return_from_code,
            "BN": "B001",
            "Expiry Date": pd.Timestamp("2027-08-01"),
            "Expiry Month Key": "2027-08",
            "Generic Item Number": "1001",
            "PackageSize": package_size,
            "Received Quantity Each": quantity_each,
            "First Received Date": pd.Timestamp("2026-08-01"),
        }
    ])


class ReturnCancelTests(unittest.TestCase):
    def test_trk_classification_is_unchanged(self):
        self.assertEqual(classify_inbound_shipment("TRK30-1"), CUSTOMER_RETURN)
        self.assertEqual(classify_inbound_shipment("TRK49-1"), STO_RETURN)

    def test_customer_return_matches_supplier_name_to_address(self):
        result = engine().build_return_cancel_reconciliation(
            returns_history(), customer_history()
        )
        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "To Address"], "Customer A - Riyadh")
        self.assertEqual(result.loc[0, "GLN"], "6281234567890")
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 3)
        self.assertEqual(
            result.loc[0, "Return Reconciliation Status"],
            "Cancel Dispatch Required",
        )

    def test_dispatch_normalizer_preserves_ship_to_customer_as_code(self):
        raw = pd.DataFrame([{
            "Batch/Lot": "B001",
            "Best Before Date": "2027-08-31",
            "Pick Qty": 30,
            "Trade Description": "Drug A",
            "To Address": "Customer A",
            "Ship To Customer": "W2D5_0402",
            "Sales Order Number": "SO1",
            "Confirm Date": "2026-08-01 10:00:00",
            "Generic Item Number": "1001",
        }])
        normalized = Normalizer.normalize_dispatch(raw)
        self.assertEqual(normalized.loc[0, "Customer Code"], "W2D5_0402")

    def test_customer_code_does_not_change_dispatch_event_identity(self):
        base = pd.DataFrame([{
            "BN": "B001",
            "Expiry Month Key": "2027-08",
            "Expiry Date": pd.Timestamp("2027-08-31"),
            "Generic Item Number": "1001",
            "Trade Item Number": "1001",
            "Trade Name": "Drug A",
            "Dispatched Quantity": 30,
            "To Address": "Customer A",
            "Customer Code": "120400",
            "Sales Order Number": "SO1",
            "Order Line": "1",
            "Dispatch Date": pd.Timestamp("2026-08-01 10:00:00"),
            "Custody": "",
            "_Source File": "dispatch.xlsx",
        }])
        subject = engine()
        subject.dispatch = base.copy()
        first_key = subject._dispatch_events().loc[0, "Event Key"]
        subject.dispatch.loc[0, "Customer Code"] = "UPDATED-CODE"
        second_key = subject._dispatch_events().loc[0, "Event Key"]
        self.assertEqual(first_key, second_key)

    def test_customer_code_matches_when_supplier_and_customer_names_differ(self):
        result = engine().build_return_cancel_reconciliation(
            returns_history(
                return_from="WMS RETURN SUPPLIER NAME",
                return_from_code="120400",
            ),
            customer_history(customer_code="120400"),
        )
        self.assertEqual(result.loc[0, "To Address"], "Customer A - Riyadh")
        self.assertEqual(result.loc[0, "Customer Code"], "120400")
        self.assertEqual(result.loc[0, "Customer Match Method"], "CUSTOMER CODE")
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 3)

    def test_ambiguous_customer_code_never_generates_cancel_dispatch(self):
        first = customer_history(customer_code="120400")
        second = customer_history(customer_code="120400")
        second["To Address"] = "Another Customer Alias"
        second["GLN"] = "6281234567891"
        result = engine().build_return_cancel_reconciliation(
            returns_history(
                return_from="WMS RETURN SUPPLIER NAME",
                return_from_code="120400",
            ),
            pd.concat([first, second], ignore_index=True),
        )
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 0)
        self.assertEqual(
            result.loc[0, "Return Reconciliation Status"],
            "Exception - Ambiguous Customer Code",
        )
        self.assertEqual(
            result.loc[0, "Customer Match Method"],
            "AMBIGUOUS CUSTOMER CODE",
        )

    def test_customer_code_combines_name_aliases_with_one_gln(self):
        first = customer_history(quantity_each=60, customer_code="120400")
        second = customer_history(quantity_each=40, customer_code="120400")
        second["To Address"] = "Customer A Renamed"
        result = engine().build_return_cancel_reconciliation(
            returns_history(
                quantity_each=100,
                return_from="Different ASN Supplier Name",
                return_from_code="120400",
            ),
            pd.concat([first, second], ignore_index=True),
        )
        self.assertEqual(result.loc[0, "Historical Dispatch Quantity Each"], 100)
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 10)
        self.assertEqual(result.loc[0, "Customer Match Method"], "CUSTOMER CODE")

    def test_confirmed_cancel_is_not_generated_again(self):
        confirmed = pd.DataFrame([
            {
                "Return Type": CUSTOMER_RETURN,
                "BN": "B001",
                "Expiry Month Key": "2027-08",
                "Generic Item Number": "1001",
                "To Address": "Customer A - Riyadh",
                "GLN": "6281234567890",
                "Previously Confirmed Cancel Dispatch Each": 20,
                "Previously Confirmed Cancel Dispatch Pack": 2,
            }
        ])
        result = engine().build_return_cancel_reconciliation(
            returns_history(), customer_history(), confirmed
        )
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 1)

    def test_cancel_is_capped_by_original_dispatch(self):
        result = engine().build_return_cancel_reconciliation(
            returns_history(quantity_each=150), customer_history()
        )
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 10)
        self.assertIn(
            "Return Exceeds Available Historical Dispatch",
            result.loc[0, "Return Reconciliation Status"],
        )

    def test_unmatched_return_becomes_exception(self):
        result = engine().build_return_cancel_reconciliation(
            returns_history(return_from="Unknown Customer"), customer_history()
        )
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 0)
        self.assertEqual(result.loc[0, "PackageSize"], 10)
        self.assertEqual(
            result.loc[0, "Return Reconciliation Status"],
            "Exception - No Matching Customer Dispatch",
        )

    def test_unmatched_return_package_size_uses_positive_fallback(self):
        result = engine().build_return_cancel_reconciliation(
            returns_history(
                return_from="Unknown Customer",
                package_size=0,
            ),
            customer_history(),
        )
        self.assertEqual(result.loc[0, "PackageSize"], 1)
        self.assertEqual(result.loc[0, "To Be Cancel Dispatch"], 0)
        self.assertEqual(
            result.loc[0, "Return Reconciliation Status"],
            "Exception - No Matching Customer Dispatch",
        )

    def test_return_calculation_does_not_mutate_customer_history(self):
        customer = customer_history()
        before = customer.copy(deep=True)
        engine().build_return_cancel_reconciliation(returns_history(), customer)
        pd.testing.assert_frame_equal(customer, before)

    def test_sto_and_customer_returns_share_the_original_dispatch_cap(self):
        returns = pd.concat([
            returns_history(quantity_each=70, return_type=STO_RETURN),
            returns_history(quantity_each=70, return_type=CUSTOMER_RETURN),
        ], ignore_index=True)
        result = engine().build_return_cancel_reconciliation(
            returns, customer_history(quantity_each=100)
        )
        self.assertEqual(result["To Be Cancel Dispatch"].sum(), 10)

    def test_return_side_stream_does_not_change_dispatch_details(self):
        subject = engine()
        original_dispatch = pd.DataFrame([
            {"To Be Dispatch": 4, "BN": "B001", "To Address": "Customer A"}
        ])
        cancel_details = pd.DataFrame(columns=subject.RETURN_CANCEL_RECONCILIATION_COLUMNS)
        with (
            patch.object(FullReconciliationEngine, "prepare_stage2_sfda_identity", return_value=pd.DataFrame()),
            patch.object(FullReconciliationEngine, "build_dispatch_reconciliation", return_value=original_dispatch.copy()),
            patch.object(FullReconciliationEngine, "build_return_cancel_reconciliation", return_value=cancel_details),
            patch("engine.full_reconciliation.Validator.validate"),
            patch("engine.full_reconciliation.Normalizer.normalize_sfda", return_value=pd.DataFrame()),
        ):
            result = subject.run_dispatch_reconciliation(
                pd.DataFrame(),
                pd.DataFrame(),
                customer_history(),
                returns_history_df=returns_history(),
            )
        pd.testing.assert_frame_equal(result["dispatch_details"], original_dispatch)

    def test_existing_dispatch_engine_is_stable_after_return_calculation(self):
        subject = engine()
        inventory = pd.DataFrame([
            {
                "BN": "B001",
                "Expiry Date": "2027-08-31",
                "Available Quantity": 50,
                "Trade Name": "Drug A",
                "Generic Item Number": "1001",
            }
        ])
        validated_sfda = pd.DataFrame([
            {
                "GTIN": "06281234567890",
                "Drug Name": "Drug A",
                "BN": "B001",
                "Expiry Date": pd.Timestamp("2027-08-31"),
                "Expiry Month Key": "2027-08",
                "Generic Item Number": "1001",
                "Quantity": 100,
                "Active": 8,
                "Quantity sent pending": 0,
                "Quantity Receive Pending": 0,
            }
        ])
        customer = customer_history()
        before = subject.build_dispatch_reconciliation(
            inventory,
            pd.DataFrame(),
            customer,
            validated_sfda_identity_df=validated_sfda,
        )
        subject.build_return_cancel_reconciliation(returns_history(), customer)
        after = subject.build_dispatch_reconciliation(
            inventory,
            pd.DataFrame(),
            customer,
            validated_sfda_identity_df=validated_sfda,
        )
        pd.testing.assert_frame_equal(before, after)

    def test_historical_workbook_has_one_unified_returns_sheet(self):
        returns = pd.concat([
            returns_history(return_type=CUSTOMER_RETURN),
            returns_history(return_type=STO_RETURN),
        ], ignore_index=True)
        generated = Exporter.build_historical_database_workbook(
            batch_master=pd.DataFrame(),
            supplier_history=pd.DataFrame(),
            sto_incoming_history=pd.DataFrame(),
            customer_history=pd.DataFrame(),
            returns_history=returns,
        )
        payload = generated["Historical_Database.xlsx"]
        workbook = load_workbook(
            io.BytesIO(base64.b64decode(payload["content"])), read_only=True
        )
        self.assertIn("Returns History", workbook.sheetnames)
        self.assertNotIn("STO Return Cancel", workbook.sheetnames)
        self.assertEqual(len(workbook.sheetnames), 7)

    def test_cancel_csv_is_separate_and_uses_cancel_quantity(self):
        row = engine().build_return_cancel_reconciliation(
            returns_history(), customer_history()
        )
        files = Exporter.build_cancel_dispatch_files_by_customer(row)
        self.assertEqual(list(files), ["Cancel_Dispatch_6281234567890_001.csv"])
        self.assertIn("06281234567890;3;B001;31-08-2027", next(iter(files.values())))

    def test_full_dispatch_workbook_contains_cancel_dispatch_sheet(self):
        dispatch = pd.DataFrame(columns=Exporter.FULL_DISPATCH_RECONCILIATION_COLUMNS)
        cancel = engine().build_return_cancel_reconciliation(
            returns_history(), customer_history()
        )
        generated = Exporter.build_full_dispatch_reconciliation_workbook(
            dispatch,
            cancel,
        )
        payload = generated["Full_Dispatch_Reconciliation.xlsx"]
        workbook = load_workbook(
            io.BytesIO(base64.b64decode(payload["content"])),
            read_only=True,
        )
        self.assertEqual(workbook.sheetnames, ["Full Dispatch", "Cancel Dispatch"])
        cancel_sheet = workbook["Cancel Dispatch"]
        headers = [cell.value for cell in cancel_sheet[3]]
        self.assertIn("To Be Cancel Dispatch", headers)


if __name__ == "__main__":
    unittest.main()
