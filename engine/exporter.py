import base64
import io
import re
from decimal import Decimal, InvalidOperation, ROUND_FLOOR

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class Exporter:
    MAX_ROWS_PER_FILE = 20
    MAX_QUANTITY_PER_FILE = 100000
    GTIN_LENGTH = 14
    DUMMY_GLN = "99999999999999"

    BATCH_MASTER_COLUMNS = [
        "GTIN",
        "Drug Name",
        "BN",
        "Expiry Date",
        "PackageSize",
        "Quantity",
        "Active",
        "Quantity sent pending",
        "Quantity Receive Pending",
        "Generic Item Number",
        "Description",
        "Trade Description",
        "Received Quantity Each",
        "Received Quantity Pack",
        "First Received Date",
        "Last Received Date",
        "Total Dispatched Qty",
        "Total Dispatched Qty Pack",
        "First Dispatch Date",
        "Last Dispatch Date",
        "Generic Exists in SFDA",
        "Last Updated",
        "Item Family Group",
        "Custody",
    ]

    SUPPLIER_HISTORY_COLUMNS = [
        "Supplier Name", "Supplier Code", "GTIN", "Drug Name",
        "Generic Item Number", "Description", "Trade Description", "BN",
        "Expiry Date", "PackageSize", "Received Quantity Each",
        "Received Quantity Pack", "First Received Date", "Last Received Date",
        "Item Family Group",
    ]

    CUSTOMER_HISTORY_COLUMNS = [
        "To Address", "GLN", "GTIN", "Drug Name", "Generic Item Number",
        "Trade Description", "BN", "Expiry Date", "PackageSize",
        "Dispatch Quantity Each", "Dispatch Quantity Pack",
        "First Dispatch Date", "Last Dispatch Date", "Custody",
    ]

    FULL_ACCEPT_RECONCILIATION_COLUMNS = [
        "GTIN",
        "Drug Name",
        "BN",
        "Expiry Date",
        "Expiry Month Key",
        "Generic Item Number",
        "PackageSize",
        "Historical Received Quantity Each",
        "Historical Received Quantity Pack",
        "SFDA Quantity",
        "SFDA Active",
        "Quantity Sent Pending",
        "Quantity Receive Pending",
        "SFDA Confirmed Accept",
        "To Be Accept",
        "Reconciliation Status",
    ]

    ACCEPT_DISTRIBUTION_COLUMNS = [
        "GTIN",
        "Drug Name",
        "BN",
        "Expiry Date",
        "Expiry Month Key",
        "Generic Item Number",
        "SFDA Quantity",
        "Quantity Receive Pending",
        "SFDA Confirmed Accept",
        "Supplier Accept Qty",
        "STO Accept Qty",
        "Total To Be Accept",
        "Accept Source",
    ]

    @staticmethod
    def build_full_accept_reconciliation_workbook(
        accept_details,
        accept_distribution,
        file_name="Full_Accept_Reconciliation.xlsx",
    ):
        """Build Full Accept details plus the exact Supplier/STO CSV allocation."""

        generated = Exporter.build_formatted_excel_file(
            df=accept_details,
            file_name=file_name,
            sheet_name="Full Accept",
            title="One-Time Full Reconciliation - Accept",
            columns=list(accept_details.columns),
            sort_columns=["Generic Item Number", "BN", "Expiry Date"],
        )
        payload = generated[file_name]
        workbook = load_workbook(io.BytesIO(base64.b64decode(payload["content"])))
        worksheet = workbook.create_sheet(title="Accept Distribution")

        report = (
            accept_distribution.copy()
            if accept_distribution is not None
            else pd.DataFrame()
        ).reindex(columns=Exporter.ACCEPT_DISTRIBUTION_COLUMNS)
        report = report.dropna(how="all").reset_index(drop=True)

        groups = [
            (1, 6, "Batch Identification", "5B9BD5"),
            (7, 9, "SFDA State", "5B9BD5"),
            (10, 12, "Accept Allocation", "4472C4"),
            (13, 13, "Source", "70AD47"),
        ]
        for start, end, label, color in groups:
            if start < end:
                worksheet.merge_cells(
                    start_row=1, start_column=start, end_row=1, end_column=end
                )
            cell = worksheet.cell(row=1, column=start, value=label)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index in range(start, end + 1):
                worksheet.cell(row=1, column=column_index).fill = PatternFill(
                    fill_type="solid", fgColor=color
                )

        border = Border(
            left=Side(style="thin", color="B8C4CE"),
            right=Side(style="thin", color="B8C4CE"),
            top=Side(style="thin", color="B8C4CE"),
            bottom=Side(style="thin", color="B8C4CE"),
        )
        identifier_columns = {"GTIN", "BN", "Generic Item Number"}
        date_columns = {"Expiry Date"}
        max_lengths = [len(column) for column in report.columns]

        for column_index, column_name in enumerate(report.columns, start=1):
            cell = worksheet.cell(row=2, column=column_index, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=("17365D" if column_index <= 9 else "2F5597" if column_index <= 12 else "548235"),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        stripe_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
        for row_index, values in enumerate(report.itertuples(index=False, name=None), start=3):
            for offset, value in enumerate(values):
                column_name = report.columns[offset]
                if pd.isna(value):
                    value = None
                elif column_name in date_columns:
                    parsed = pd.to_datetime(value, errors="coerce")
                    if not pd.isna(parsed):
                        value = parsed.to_pydatetime()
                elif column_name in identifier_columns:
                    value = Exporter._normalize_identifier(value)
                elif hasattr(value, "item"):
                    value = value.item()
                if value is not None:
                    max_lengths[offset] = max(max_lengths[offset], len(str(value)))
                cell = worksheet.cell(row=row_index, column=offset + 1, value=value)
                cell.border = border
                if column_name in identifier_columns:
                    cell.number_format = "@"
                elif column_name in date_columns:
                    cell.number_format = "dd-mm-yyyy"
                elif isinstance(value, (int, float)):
                    cell.number_format = "#,##0.##"
                if row_index % 2 == 0:
                    cell.fill = stripe_fill

        if len(report) > 0:
            worksheet.auto_filter.ref = f"A2:M{2 + len(report)}"
        worksheet.freeze_panes = "A3"
        worksheet.sheet_view.showGridLines = False
        for column_index, max_length in enumerate(max_lengths, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 12), 45
            )

        output = io.BytesIO()
        workbook.save(output)
        payload["content"] = base64.b64encode(output.getvalue()).decode("ascii")
        return generated

    FULL_DISPATCH_RECONCILIATION_COLUMNS = [
        "To Address",
        "GLN",
        "GTIN",
        "Drug Name",
        "BN",
        "Expiry Date",
        "Expiry Month Key",
        "Generic Item Number",
        "Custody",
        "PackageSize",
        "Historical Dispatch Quantity Each",
        "Historical Dispatch Quantity Pack",
        "Previously Confirmed Full Dispatch Each",
        "Previously Confirmed Full Dispatch Pack",
        "Reserved Full Dispatch Quantity Each",
        "Reserved Full Dispatch Quantity Pack",
        "Available Historical Dispatch Quantity Each",
        "Available Historical Dispatch Quantity Pack",
        "Current Inventory Quantity Each",
        "Current Inventory Quantity Pack",
        "SFDA Quantity",
        "SFDA Active",
        "Quantity Sent Pending",
        "Quantity Receive Pending",
        "To Be Dispatch",
        "Reconciliation Status",
    ]

    SUPPLIER_VARIANCE_COLUMNS = [
        "Supplier Name",
        "Supplier Code",
        "GTIN",
        "Drug Name",
        "BN",
        "Expiry Date",
        "Expiry Month Key",
        "Generic Item Number",
        "Historical Received Quantity Each",
        "Historical Received Quantity Pack",
        "SFDA Supplier Quantity",
        "Supplier Variance",
        "Variance Status",
        "Required Action",
    ]

    ACCEPT_DETAILS_COLUMNS = [
        "GTIN",
        "Drug Name",
        "BN",
        "Expiry Date",
        "PackageSize",
        "Active",
        "Quantity sent pending",
        "Quantity Receive Pending",
        "Generic Item Number",
        "Received Quantity Each",
        "Received Quantity Pack",
        "Description",
        "Inbound Shipment",
        "Supplier Name",
        "Supplier Code",
        "Item Family Group",
        "To Be Accept",
        "Processing Status",
        "Previous Run Date",
        "Previous Quantity Each",
        "Current Quantity Each",
        "Quantity Difference",
        "Package Size Status",
        "Batch Master Status",
    ]

    DISPATCH_DETAILS_COLUMNS = [
        "GTIN",
        "Drug Name",
        "BN",
        "Expiry Date",
        "PackageSize",
        "Active",
        "Quantity sent pending",
        "Quantity Receive Pending",
        "Generic Item Number",
        "Trade Name",
        "Sales Order Number",
        "Order Line",
        "To Address",
        "Dispatch Date",
        "Dispatch Quantity Each",
        "Dispatch Quantity Pack",
        "Allocated To Be Dispatch",
        "Processing Status",
        "Previous Run Date",
        "Previous Quantity Each",
        "Current Quantity Each",
        "Quantity Difference",
        "GLN",
        "Customer Status",
        "Package Size Status",
        "Batch Master Status",
    ]

    @staticmethod
    def _normalize_identifier(value):
        if pd.isna(value):
            return ""

        text = str(value).strip()

        if text.endswith(".0"):
            text = text[:-2]

        if re.fullmatch(
            r"[+-]?\d+(\.\d+)?[eE][+-]?\d+",
            text,
        ):
            try:
                text = (
                    format(Decimal(text), "f")
                    .rstrip("0")
                    .rstrip(".")
                )
            except InvalidOperation:
                pass

        return text

    @staticmethod
    def _normalize_gtin(value):
        gtin = (
            Exporter._normalize_identifier(value)
            .replace(" ", "")
        )

        if gtin.isdigit():
            gtin = gtin.zfill(
                Exporter.GTIN_LENGTH
            )

        return gtin

    @staticmethod
    def _normalize_quantity(value):
        quantity = pd.to_numeric(
            value,
            errors="coerce",
        )

        if pd.isna(quantity):
            return 0

        try:
            return max(
                0,
                int(
                    Decimal(str(quantity))
                    .to_integral_value(
                        rounding=ROUND_FLOOR
                    )
                ),
            )
        except (
            InvalidOperation,
            ValueError,
            TypeError,
        ):
            return 0

    @staticmethod
    def _normalize_expiry(value):
        expiry = pd.to_datetime(
            value,
            errors="coerce",
            dayfirst=True,
        )

        if pd.isna(expiry):
            return ""

        return expiry.strftime("%d-%m-%Y")

    @staticmethod
    def _safe_file_name(value):
        text = re.sub(
            r"[^A-Za-z0-9_-]+",
            "_",
            str(value).strip(),
        )

        return (
            text.strip("_")
            or Exporter.DUMMY_GLN
        )

    @staticmethod
    def _prepare_records(df, quantity_column):
        records = []

        if df is None or df.empty:
            return records

        for _, row in df.iterrows():
            quantity = Exporter._normalize_quantity(
                row.get(quantity_column)
            )

            if quantity <= 0:
                continue

            gtin = Exporter._normalize_gtin(
                row.get("GTIN")
            )
            batch = str(
                row.get("BN", "")
            ).strip()
            expiry = Exporter._normalize_expiry(
                row.get("Expiry Date")
            )

            if not gtin or not batch or not expiry:
                continue

            records.append({
                "GTIN": gtin,
                "QUANTITY": quantity,
                "BN": batch,
                "XD": expiry,
            })

        return records

    @staticmethod
    def _aggregate_dispatch_records(records):
        """Aggregate identical SFDA batch rows before Dummy-GLN file splitting.

        Only the final regulatory identity is relevant to SFDA here:
            GTIN + BN + exact SFDA expiry date (XD).

        This is intentionally used only for the Dummy GLN pool. Real GLN
        exports preserve the existing customer-level behavior.
        """
        if not records:
            return []

        aggregated = {}
        order = []

        for record in records:
            key = (
                str(record.get("GTIN", "")).strip(),
                str(record.get("BN", "")).strip(),
                str(record.get("XD", "")).strip(),
            )

            if key not in aggregated:
                aggregated[key] = {
                    "GTIN": key[0],
                    "QUANTITY": 0,
                    "BN": key[1],
                    "XD": key[2],
                }
                order.append(key)

            aggregated[key]["QUANTITY"] += int(
                record.get("QUANTITY", 0) or 0
            )

        return [
            aggregated[key]
            for key in order
            if int(aggregated[key]["QUANTITY"]) > 0
        ]

    @staticmethod
    def _split_into_files(records):
        files = []
        current_rows = []
        current_quantity = 0

        for record in records:
            remaining = int(record["QUANTITY"])

            while remaining > 0:
                if (
                    len(current_rows)
                    >= Exporter.MAX_ROWS_PER_FILE
                    or current_quantity
                    >= Exporter.MAX_QUANTITY_PER_FILE
                ):
                    files.append(current_rows)
                    current_rows = []
                    current_quantity = 0

                available = (
                    Exporter.MAX_QUANTITY_PER_FILE
                    - current_quantity
                )

                quantity = min(
                    remaining,
                    available,
                )

                new_record = record.copy()
                new_record["QUANTITY"] = quantity

                current_rows.append(new_record)
                current_quantity += quantity
                remaining -= quantity

        if current_rows:
            files.append(current_rows)

        return files

    @staticmethod
    def _build_sfda_content(records):
        lines = ["GTIN;QUANTITY;BN;XD"]

        for record in records:
            lines.append(
                "{};{};{};{}".format(
                    record["GTIN"],
                    int(record["QUANTITY"]),
                    record["BN"],
                    record["XD"],
                )
            )

        return "\r\n".join(lines)

    @staticmethod
    def build_sfda_upload_files(
        df,
        quantity_column,
        file_prefix,
    ):
        groups = Exporter._split_into_files(
            Exporter._prepare_records(
                df,
                quantity_column,
            )
        )

        output = {}

        for index, records_group in enumerate(
            groups,
            start=1,
        ):
            file_name = (
                f"{file_prefix}_{index:03d}.csv"
            )
            output[file_name] = (
                Exporter._build_sfda_content(
                    records_group
                )
            )

        return output

    @staticmethod
    def build_dispatch_files_by_customer(
        dispatch_df,
    ):
        output = {}

        if dispatch_df is None or dispatch_df.empty:
            return output

        working = dispatch_df.copy()
        if "GLN" not in working.columns:
            working["GLN"] = ""

        working["Export GLN"] = working["GLN"].apply(
            Exporter._normalize_identifier
        )

        # Missing GLN rows belong to ONE regulatory Dummy-GLN pool.
        # Customer / To Address remains available in Dispatch Details, but it
        # must not create separate SFDA upload files.
        missing_gln = (
            working["Export GLN"].eq("")
            | working["Export GLN"].str.upper().eq("DUMMY")
        )
        working.loc[missing_gln, "Export GLN"] = Exporter.DUMMY_GLN

        quantity_column = (
            "To Be Dispatch"
            if "To Be Dispatch" in working.columns
            else "Allocated To Be Dispatch"
        )

        for customer_gln, customer_df in working.groupby(
            "Export GLN",
            dropna=False,
            sort=True,
        ):
            safe_gln = Exporter._safe_file_name(customer_gln)

            records = Exporter._prepare_records(
                customer_df,
                quantity_column,
            )

            # Dummy GLN:
            #   1) consolidate all customers into one pool;
            #   2) combine the same GTIN + BN + exact SFDA expiry;
            #   3) split only by the regulatory limits:
            #        - max 20 rows per CSV
            #        - max 100,000 total quantity per CSV.
            #
            # Real GLNs intentionally keep the existing behavior unchanged.
            normalized_gln = Exporter._normalize_identifier(customer_gln)
            is_dummy_gln = (
                normalized_gln == Exporter.DUMMY_GLN
                or (
                    normalized_gln != ""
                    and set(normalized_gln) == {"9"}
                    and len(normalized_gln) in {13, 14}
                )
            )
            if is_dummy_gln:
                records = Exporter._aggregate_dispatch_records(records)

            groups = Exporter._split_into_files(records)

            for index, records_group in enumerate(groups, start=1):
                file_name = f"{safe_gln}_{index:03d}.csv"
                output[file_name] = Exporter._build_sfda_content(
                    records_group
                )

        return output

    @staticmethod
    def _is_batch_master_report(file_name, sheet_name):
        file_text = str(file_name or "").strip().lower()
        sheet_text = str(sheet_name or "").strip().lower()
        return (
            "batch_master" in file_text
            or "batch master" in file_text
            or sheet_text == "batch master"
        )

    @staticmethod
    def _is_supplier_history_report(file_name, sheet_name):
        text = f"{file_name or ''} {sheet_name or ''}".strip().lower()
        return "supplier_history" in text or "supplier history" in text

    @staticmethod
    def _is_customer_history_report(file_name, sheet_name):
        text = f"{file_name or ''} {sheet_name or ''}".strip().lower()
        return "customer_history" in text or "customer history" in text

    @staticmethod
    def _is_full_accept_reconciliation_report(file_name, sheet_name):
        text = f"{file_name or ''} {sheet_name or ''}".strip().lower()
        return (
            "full_accept_reconciliation" in text
            or "full accept reconciliation" in text
        )

    @staticmethod
    def _is_full_dispatch_reconciliation_report(file_name, sheet_name):
        text = f"{file_name or ''} {sheet_name or ''}".strip().lower()
        return (
            "full_dispatch_reconciliation" in text
            or "full dispatch reconciliation" in text
        )

    @staticmethod
    def _is_supplier_variance_report(file_name, sheet_name):
        text = f"{file_name or ''} {sheet_name or ''}".strip().lower()
        return (
            "supplier_variance" in text
            or "supplier variance" in text
        )


    @staticmethod
    def _is_full_reconciliation_summary_report(file_name, sheet_name):
        text = f"{file_name or ''} {sheet_name or ''}".strip().lower()
        return (
            "full_reconciliation_summary" in text
            or "full reconciliation summary" in text
        )

    @staticmethod
    def _is_accept_details_report(file_name, sheet_name):
        file_text = str(file_name or "").strip().lower()
        sheet_text = str(sheet_name or "").strip().lower()

        return (
            "accept_details" in file_text
            or "accept details" in file_text
            or sheet_text == "accept details"
        )

    @staticmethod
    def _is_dispatch_details_report(file_name, sheet_name):
        file_text = str(file_name or "").strip().lower()
        sheet_text = str(sheet_name or "").strip().lower()

        return (
            "dispatch_details" in file_text
            or "dispatch details" in file_text
            or sheet_text == "dispatch details"
        )

    @staticmethod
    def build_batch_master_two_sheet_file(
        df,
        file_name="Batch_Master.xlsx",
    ):
        """Build Batch Master workbook with two explicit business stages.

        Sheet 1 - Matched SFDA Batches:
            direct BN + Expiry Month matches only (Generic Exists in SFDA = Yes).

        Sheet 2 - Missing From SFDA:
            WMS batches belonging to a Generic proven in Sheet 1, excluding all
            batches already present in Sheet 1.
        """
        source = df.copy() if df is not None else pd.DataFrame()
        source = source.reindex(columns=Exporter.BATCH_MASTER_COLUMNS)
        status = source.get(
            "Generic Exists in SFDA",
            pd.Series("", index=source.index, dtype=object),
        ).fillna("").astype(str).str.strip().str.upper()

        matched = source.loc[status.eq("YES")].copy()
        missing = source.loc[status.eq("MISSING BATCH IN SFDA")].copy()

        sort_columns = [
            column for column in ["Generic Item Number", "BN", "Expiry Date"]
            if column in source.columns
        ]
        if sort_columns:
            matched = matched.sort_values(sort_columns, kind="stable")
            missing = missing.sort_values(sort_columns, kind="stable")

        workbook = Workbook()
        workbook.remove(workbook.active)

        border = Border(
            left=Side(style="thin", color="B8C4CE"),
            right=Side(style="thin", color="B8C4CE"),
            top=Side(style="thin", color="B8C4CE"),
            bottom=Side(style="thin", color="B8C4CE"),
        )

        def add_sheet(frame, sheet_name, title):
            ws = workbook.create_sheet(title=sheet_name[:31])
            frame = frame.dropna(how="all").reset_index(drop=True)
            columns = list(frame.columns)
            last_column = get_column_letter(max(1, len(columns)))

            ws.merge_cells(f"A1:{last_column}1")
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
            ws["A1"].fill = PatternFill(fill_type="solid", fgColor="0F6CBD")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            # Same SFDA/WMS visual split used by the normal Batch Master report.
            group_definitions = [
                (1, 9, "SFDA Report", "5B9BD5"),
                (10, len(columns), "WMS Report", "4472C4"),
            ]
            for start_col, end_col, label, fill_color in group_definitions:
                if start_col > len(columns) or end_col < start_col:
                    continue
                end_col = min(end_col, len(columns))
                ws.merge_cells(
                    f"{get_column_letter(start_col)}2:{get_column_letter(end_col)}2"
                )
                cell = ws.cell(row=2, column=start_col, value=label)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                for col in range(start_col, end_col + 1):
                    ws.cell(row=2, column=col).fill = PatternFill(
                        fill_type="solid", fgColor=fill_color
                    )

            header_row = 3
            identifier_flags = []
            date_flags = []
            max_lengths = []
            for col_idx, name in enumerate(columns, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=str(name))
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="17365D" if col_idx <= 9 else "2F5597",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                lower = str(name).lower()
                identifier_flags.append(any(x in lower for x in [
                    "gtin", "generic item", "trade item", "gln", "bn"
                ]))
                date_flags.append("date" in lower or "expiry" in lower)
                max_lengths.append(len(str(name)))

            stripe_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
            for row_offset, values in enumerate(frame.itertuples(index=False, name=None), start=1):
                excel_row = header_row + row_offset
                for idx, value in enumerate(values):
                    col_idx = idx + 1
                    if pd.isna(value):
                        value = None
                    elif date_flags[idx]:
                        parsed = pd.to_datetime(value, errors="coerce")
                        if not pd.isna(parsed):
                            value = parsed.to_pydatetime()
                    elif identifier_flags[idx]:
                        value = Exporter._normalize_identifier(value)
                    elif hasattr(value, "item"):
                        try:
                            value = value.item()
                        except Exception:
                            pass
                    if value is not None:
                        max_lengths[idx] = max(max_lengths[idx], len(str(value)))
                    cell = ws.cell(row=excel_row, column=col_idx, value=value)
                    cell.border = border
                    if identifier_flags[idx]:
                        cell.number_format = "@"
                    elif date_flags[idx]:
                        cell.number_format = "dd-mm-yyyy"
                    elif isinstance(value, (int, float)):
                        cell.number_format = "#,##0.##"
                    if row_offset % 2 == 0:
                        cell.fill = stripe_fill

            if len(frame) > 0:
                ws.auto_filter.ref = f"A{header_row}:{last_column}{header_row + len(frame)}"
            ws.freeze_panes = f"A{header_row + 1}"
            ws.sheet_view.showGridLines = False
            for idx, max_length in enumerate(max_lengths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = min(
                    max(max_length + 2, 12), 45
                )

        add_sheet(
            matched,
            "Matched SFDA Batches",
            "Stage 1 - Direct BN + Expiry Month Matches",
        )
        add_sheet(
            missing,
            "Missing From SFDA",
            "Stage 2 - Trusted Generic Batches Missing From SFDA",
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        content = base64.b64encode(output.read()).decode("ascii")
        return {
            file_name: {
                "content": content,
                "encoding": "base64",
                "mime_type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            }
        }


    @staticmethod
    def build_historical_database_workbook(
        batch_master,
        supplier_history,
        sto_incoming_history,
        customer_history,
        sto_return_history,
        file_name="Historical_Database.xlsx",
    ):
        """Build the complete Historical Database as one seven-sheet workbook.

        Business contract:
        - Batch Master is batch movement history only; Supplier Name/Code are not exported.
        - Batch Master contains direct SFDA matches (BN + Expiry Month).
        - Missing From SFDA contains received WMS batches for Generics proven by Stage 1.
        - Supplier/STO/Customer histories remain separate audit views in their own sheets.
        """
        master = batch_master.copy() if batch_master is not None else pd.DataFrame()
        master = master.reindex(columns=Exporter.BATCH_MASTER_COLUMNS)
        status = master.get(
            "Generic Exists in SFDA",
            pd.Series("", index=master.index, dtype=object),
        ).fillna("").astype(str).str.strip().str.upper()
        matched = master.loc[status.eq("YES")].copy()
        missing = master.loc[status.eq("MISSING BATCH IN SFDA")].copy()

        def prepared(frame, columns=None, sort_columns=None):
            out = frame.copy() if frame is not None else pd.DataFrame()
            if columns is not None:
                out = out.reindex(columns=columns)
            if sort_columns:
                valid = [c for c in sort_columns if c in out.columns]
                if valid:
                    out = out.sort_values(valid, kind="stable")
            return out.dropna(how="all").reset_index(drop=True)

        matched = prepared(matched, sort_columns=["Generic Item Number", "BN", "Expiry Date"])
        missing = prepared(missing, sort_columns=["Generic Item Number", "BN", "Expiry Date"])
        supplier = prepared(
            supplier_history,
            Exporter.SUPPLIER_HISTORY_COLUMNS,
            ["Supplier Name", "Generic Item Number", "BN", "Expiry Date"],
        )
        sto_in = prepared(
            sto_incoming_history,
            sort_columns=["Source Warehouse", "Generic Item Number", "BN", "Expiry Date"],
        )
        customer = prepared(
            customer_history,
            Exporter.CUSTOMER_HISTORY_COLUMNS,
            ["To Address", "Generic Item Number", "BN", "Expiry Date"],
        )
        sto_return = prepared(
            sto_return_history,
            sort_columns=["Source Warehouse", "Generic Item Number", "BN", "Expiry Date"],
        )

        workbook = Workbook()
        workbook.remove(workbook.active)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill = PatternFill(fill_type="solid", fgColor="0F6CBD")
        header_fill = PatternFill(fill_type="solid", fgColor="17365D")
        alt_fill = PatternFill(fill_type="solid", fgColor="F4F8FC")

        def add_data_sheet(frame, sheet_name, title):
            ws = workbook.create_sheet(title=sheet_name[:31])
            cols = list(frame.columns)
            last_col = get_column_letter(max(1, len(cols)))
            ws.merge_cells(f"A1:{last_col}1")
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
            ws["A1"].fill = title_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            header_row = 3
            max_lengths = [len(str(c)) for c in cols]
            for idx, col in enumerate(cols, start=1):
                cell = ws.cell(header_row, idx, str(col))
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for r_idx, values in enumerate(frame.itertuples(index=False, name=None), start=header_row + 1):
                for c_idx, value in enumerate(values, start=1):
                    if pd.isna(value):
                        value = ""
                    cell = ws.cell(r_idx, c_idx, value)
                    cell.border = border
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill
                    if c_idx <= len(max_lengths):
                        max_lengths[c_idx - 1] = min(45, max(max_lengths[c_idx - 1], len(str(value))))
            if cols:
                ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(frame)}"
                ws.freeze_panes = f"A{header_row + 1}"
                for idx, width in enumerate(max_lengths, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 12), 45)
            ws.sheet_view.showGridLines = False

        # Summary is intentionally first so users can understand the workbook immediately.
        summary_ws = workbook.create_sheet(title="Summary")
        summary_ws.merge_cells("A1:C1")
        summary_ws["A1"] = "Historical Database Summary"
        summary_ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
        summary_ws["A1"].fill = title_fill
        summary_ws["A1"].alignment = Alignment(horizontal="center")
        summary_ws.append([])
        summary_ws.append(["Metric", "Value", "Description"])
        for cell in summary_ws[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border

        def qty_sum(frame, column):
            if column not in frame.columns:
                return 0
            return float(pd.to_numeric(frame[column], errors="coerce").fillna(0).sum())

        metrics = [
            ("Matched SFDA Batches", len(matched), "Direct BN + Expiry Month matches"),
            ("Missing From SFDA", len(missing), "Received batches for trusted Generics not present in SFDA"),
            ("Total Historical Batches", len(matched) + len(missing), "Matched + Missing batch records"),
            ("Supplier History Rows", len(supplier), "Supplier receipt history"),
            ("STO Incoming Rows", len(sto_in), "Incoming inter-warehouse transfer history"),
            ("Customer History Rows", len(customer), "Customer dispatch history"),
            ("STO Return/Cancel Rows", len(sto_return), "STO return / cancel-dispatch history"),
            ("Total Received Qty Each", qty_sum(master, "Received Quantity Each"), "Historical receipts across eligible receipt types"),
            ("Total Dispatched Qty Each", qty_sum(master, "Total Dispatched Qty"), "Historical dispatch quantity"),
        ]
        for metric, value, desc in metrics:
            summary_ws.append([metric, value, desc])
        for row in summary_ws.iter_rows(min_row=4, max_row=3 + len(metrics), min_col=1, max_col=3):
            for cell in row:
                cell.border = border
        summary_ws.column_dimensions["A"].width = 30
        summary_ws.column_dimensions["B"].width = 20
        summary_ws.column_dimensions["C"].width = 58
        summary_ws.freeze_panes = "A4"
        summary_ws.sheet_view.showGridLines = False

        add_data_sheet(matched, "Batch Master", "Historical Batch Master - Matched SFDA Batches")
        add_data_sheet(missing, "Missing From SFDA", "Trusted Generic Batches Missing From SFDA")
        add_data_sheet(supplier, "Supplier History", "Historical Supplier Receipt History")
        add_data_sheet(sto_in, "STO Incoming History", "Historical STO Incoming Receipt History")
        add_data_sheet(customer, "Customer History", "Historical Customer Dispatch History")
        add_data_sheet(sto_return, "STO Return Cancel", "STO Returns - Cancel Previous RSD Dispatch")

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        content = base64.b64encode(output.read()).decode("ascii")
        return {
            file_name: {
                "content": content,
                "encoding": "base64",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        }

    @staticmethod
    def build_formatted_excel_file(
        df,
        file_name,
        sheet_name,
        title,
        columns=None,
        sort_columns=None,
    ):
        report = (
            df.copy()
            if df is not None
            else pd.DataFrame()
        )

        is_batch_master = Exporter._is_batch_master_report(
            file_name,
            sheet_name,
        )
        is_supplier_history = Exporter._is_supplier_history_report(file_name, sheet_name)
        is_customer_history = Exporter._is_customer_history_report(file_name, sheet_name)
        is_full_accept_reconciliation = (
            Exporter._is_full_accept_reconciliation_report(
                file_name,
                sheet_name,
            )
        )
        is_full_dispatch_reconciliation = (
            Exporter._is_full_dispatch_reconciliation_report(
                file_name,
                sheet_name,
            )
        )
        is_supplier_variance = Exporter._is_supplier_variance_report(
            file_name,
            sheet_name,
        )
        is_full_reconciliation_summary = (
            Exporter._is_full_reconciliation_summary_report(
                file_name,
                sheet_name,
            )
        )
        is_accept_details = Exporter._is_accept_details_report(
            file_name,
            sheet_name,
        )
        is_dispatch_details = Exporter._is_dispatch_details_report(
            file_name,
            sheet_name,
        )

        if is_batch_master:
            report = report.reindex(columns=Exporter.BATCH_MASTER_COLUMNS)
        elif is_supplier_history:
            report = report.reindex(columns=Exporter.SUPPLIER_HISTORY_COLUMNS)
        elif is_customer_history:
            report = report.reindex(columns=Exporter.CUSTOMER_HISTORY_COLUMNS)
        elif is_full_accept_reconciliation:
            report = report.reindex(
                columns=Exporter.FULL_ACCEPT_RECONCILIATION_COLUMNS
            )
        elif is_full_dispatch_reconciliation:
            report = report.reindex(
                columns=Exporter.FULL_DISPATCH_RECONCILIATION_COLUMNS
            )
        elif is_supplier_variance:
            report = report.reindex(
                columns=Exporter.SUPPLIER_VARIANCE_COLUMNS
            )
        elif is_accept_details:
            report = report.reindex(
                columns=Exporter.ACCEPT_DETAILS_COLUMNS
            )
        elif is_dispatch_details:
            report = report.reindex(
                columns=Exporter.DISPATCH_DETAILS_COLUMNS
            )

        if columns and not (
            is_batch_master
            or is_accept_details
            or is_dispatch_details
            or is_full_accept_reconciliation
            or is_full_dispatch_reconciliation
            or is_supplier_variance
        ):
            report = report[
                [
                    column
                    for column in columns
                    if column in report.columns
                ]
            ]

        if sort_columns:
            available_sort_columns = [
                column
                for column in sort_columns
                if column in report.columns
            ]

            if available_sort_columns:
                report = report.sort_values(
                    by=available_sort_columns,
                    kind="stable",
                )

        report = report.dropna(how="all").reset_index(drop=True)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name[:31]

        column_count = max(
            1,
            len(report.columns),
        )
        last_column = get_column_letter(
            column_count
        )

        is_stage2_report = (
            is_full_accept_reconciliation
            or is_full_dispatch_reconciliation
            or is_supplier_variance
            or is_full_reconciliation_summary
        )

        if (
            is_batch_master
            or is_accept_details
            or is_dispatch_details
            or is_full_accept_reconciliation
            or is_full_dispatch_reconciliation
            or is_supplier_variance
        ):
            if is_batch_master:
                group_definitions = [
                    (1, 9, "SFDA Report", "5B9BD5"),
                    (10, 24, "WMS Report", "4472C4"),
                ]
            elif is_full_accept_reconciliation:
                group_definitions = [
                    (1, 7, "Batch Identification", "5B9BD5"),
                    (8, 14, "Quantity Comparison", "4472C4"),
                    (15, 16, "Decision", "70AD47"),
                ]
            elif is_full_dispatch_reconciliation:
                group_definitions = [
                    (1, 9, "Customer and Batch", "5B9BD5"),
                    (10, 23, "Quantity Comparison", "4472C4"),
                    (24, 25, "Decision", "70AD47"),
                ]
            elif is_supplier_variance:
                group_definitions = [
                    (1, 8, "Supplier and Batch", "5B9BD5"),
                    (9, 12, "Quantity Comparison", "4472C4"),
                    (13, 14, "Decision", "70AD47"),
                ]
            elif is_accept_details:
                group_definitions = [
                    (1, 8, "SFDA Report", "5B9BD5"),
                    (9, 16, "WMS Receiving Report", "4472C4"),
                    (17, 24, "Decision", "70AD47"),
                ]
            else:
                group_definitions = [
                    (1, 8, "SFDA Report", "5B9BD5"),
                    (9, 16, "WMS Dispatch Report", "4472C4"),
                    (17, 26, "Decision", "70AD47"),
                ]

            for start_column, end_column, group_title, fill_color in group_definitions:
                start_letter = get_column_letter(start_column)
                end_letter = get_column_letter(end_column)
                worksheet.merge_cells(
                    f"{start_letter}1:{end_letter}1"
                )
                group_cell = worksheet.cell(
                    row=1,
                    column=start_column,
                    value=group_title,
                )
                group_cell.font = Font(
                    bold=True,
                    color="FFFFFF",
                    size=11,
                )
                group_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=fill_color,
                )
                group_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

                for column_index in range(start_column, end_column + 1):
                    worksheet.cell(row=1, column=column_index).fill = PatternFill(
                        fill_type="solid",
                        fgColor=fill_color,
                    )

            worksheet.row_dimensions[1].height = 22
            header_row = 2
        else:
            worksheet.merge_cells(
                f"A1:{last_column}1"
            )
            worksheet["A1"] = title
            worksheet["A1"].font = Font(
                bold=True,
                color="FFFFFF",
                size=16,
            )
            worksheet["A1"].fill = PatternFill(
                fill_type="solid",
                fgColor="0F6CBD",
            )
            worksheet["A1"].alignment = Alignment(
                horizontal="center"
            )
            header_row = (
                2
                if is_full_reconciliation_summary
                else 3
            )
        border = Border(
            left=Side(
                style="thin",
                color="B8C4CE",
            ),
            right=Side(
                style="thin",
                color="B8C4CE",
            ),
            top=Side(
                style="thin",
                color="B8C4CE",
            ),
            bottom=Side(
                style="thin",
                color="B8C4CE",
            ),
        )

        for column_index, column_name in enumerate(
            report.columns,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=str(column_name),
            )
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )
            if is_batch_master:
                header_fill = "2F5597" if column_index >= 10 else "17365D"
            elif is_full_accept_reconciliation and column_index >= 15:
                header_fill = "548235"
            elif is_full_accept_reconciliation and column_index >= 8:
                header_fill = "2F5597"
            elif is_full_dispatch_reconciliation and column_index >= 18:
                header_fill = "548235"
            elif is_full_dispatch_reconciliation and column_index >= 10:
                header_fill = "2F5597"
            elif is_supplier_variance and column_index >= 13:
                header_fill = "548235"
            elif is_supplier_variance and column_index >= 9:
                header_fill = "2F5597"
            elif (
                is_accept_details
                or is_dispatch_details
            ) and column_index >= 17:
                header_fill = "548235"
            elif (
                is_accept_details
                or is_dispatch_details
            ) and column_index >= 9:
                header_fill = "2F5597"
            else:
                header_fill = "17365D"

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=header_fill,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border

        column_names = list(report.columns)
        column_texts = [str(name).lower() for name in column_names]
        identifier_flags = [
            any(identifier in text for identifier in [
                "gtin", "generic item", "trade item", "sales order",
                "order number", "order line", "inbound shipment", "gln", "bn",
            ])
            for text in column_texts
        ]
        date_flags = [("date" in text or "expiry" in text) for text in column_texts]
        max_lengths = [len(str(name)) for name in column_names]
        stripe_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")

        # itertuples avoids the per-row Series construction cost of iterrows().
        # Width statistics are collected during the same pass, eliminating a
        # second full scan of every report column.
        for row_index, values in enumerate(report.itertuples(index=False, name=None)):
            excel_row = header_row + row_index + 1
            for offset, value in enumerate(values):
                column_index = offset + 1
                if pd.isna(value):
                    value = None
                elif date_flags[offset]:
                    converted = pd.to_datetime(value, errors="coerce")
                    if not pd.isna(converted):
                        value = converted.to_pydatetime()
                elif identifier_flags[offset]:
                    value = Exporter._normalize_identifier(value)
                elif hasattr(value, "item"):
                    try:
                        value = value.item()
                    except Exception:
                        pass

                if value is not None:
                    max_lengths[offset] = max(max_lengths[offset], len(str(value)))

                cell = worksheet.cell(row=excel_row, column=column_index, value=value)
                cell.border = border
                if identifier_flags[offset]:
                    cell.number_format = "@"
                elif date_flags[offset]:
                    cell.number_format = "dd-mm-yyyy"
                elif isinstance(value, (int, float)):
                    cell.number_format = "#,##0.##"
                if row_index % 2 == 1:
                    cell.fill = stripe_fill

        if len(report) > 0:
            report_range = (
                f"A{header_row}:"
                f"{last_column}"
                f"{header_row + len(report)}"
            )

            if (
                is_batch_master
                or is_accept_details
                or is_dispatch_details
                or is_stage2_report
            ):
                worksheet.auto_filter.ref = report_range
            else:
                table = Table(
                    displayName="ReportTable",
                    ref=report_range,
                )
                table.tableStyleInfo = (
                    TableStyleInfo(
                        name="TableStyleMedium2",
                        showRowStripes=False,
                    )
                )
                worksheet.add_table(table)

        worksheet.freeze_panes = (
            f"A{header_row + 1}"
        )
        worksheet.sheet_view.showGridLines = False

        for column_index, max_length in enumerate(max_lengths, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 12), 45
            )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        content = base64.b64encode(
            output.read()
        ).decode("ascii")

        return {
            file_name: {
                "content": content,
                "encoding": "base64",
                "mime_type": (
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
            }
        }
